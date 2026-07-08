import streamlit as st
from openpyxl import load_workbook
import requests
import json
import os

# --- 1. CONFIGURATION & AI INITIALIZATION ---
# ⚠️ MAKE SURE YOUR KEY IS ACTIVE IN YOUR GOOGLE AI STUDIO CONSOLE
GEMINI_API_KEY = "AQ.Ab8RN6KsLatdsnaJYMBNelQu1vASBvX-bFxJciUC8EKLjzJzHg"

st.set_page_config(page_title="AI House Plan Consultant", page_icon="🏠", layout="wide")
st.title("🤖 AI House Plan Consultant")
st.write("Talk to our AI agent to find your ideal layout. Tell it your budget, room requirements, or family size!")
st.write("---")

# --- 2. LOAD DATA FROM EXCEL ---
@st.cache_data
def load_excel_data():
    # We will test common file patterns to find your file automatically
    possible_names = ["house_data.xlsx", "house_data.XLSX", "house_data.xls", "house_data"]
    
    filename = None
    for name in possible_names:
        if os.path.exists(name):
            filename = name
            break
            
    if not filename:
        # Debugging step: let's see what files actually exist in this directory
        try:
            current_files = os.listdir(".")
            st.sidebar.warning(f"Files found in folder: {current_files}")
        except:
            pass
        return []
        
    try:
        wb = load_workbook(filename, data_only=True)
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]
        house_list = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(row):
                house_list.append(dict(zip(headers, row)))
        return house_list
    except Exception as e:
        st.sidebar.error(f"Excel read crash: {e}")
        return []

all_houses = load_excel_data()

# --- 3. HELPER FUNCTION TO MAP IMAGES ---
def get_image_path(excel_path):
    clean_id = str(excel_path).replace("3D_Models/", "").replace("P", "").strip()
    try:
        padded_id = f"P{int(clean_id):03d}"
    except ValueError:
        padded_id = f"P{clean_id}"
        
    target_folder = f"3D_Models_{padded_id}"
    if os.path.exists(target_folder):
        for file in os.listdir(target_folder):
            if file.upper().startswith("FLOOR_") and file.lower().endswith(('.png', '.jpg', '.jpeg')):
                return os.path.join(target_folder, file)
    return None

# --- 4. INITIALIZE CHAT HISTORY ---
if "messages" not in st.session_state:
    st.session_state.messages = [
        {"role": "assistant", "content": "Hello! I am your AI architect consultant. Tell me, what kind of house layout are you looking to build, and what is your target budget?"}
    ]

# Display historical messages
for message in st.session_state.messages:
    with st.chat_message(message["role"]):
        st.write(message["content"])
        if "matched_plans" in message:
            for house in message["matched_plans"]:
                with st.expander(f"📋 View Blueprint Details for {house.get('3D_Folder_Path')}"):
                    c1, c2 = st.columns([1, 1])
                    with c1:
                        st.write(f"**🛏️ Bedrooms:** {house.get('NO OF BEDROOM', '-')}")
                        st.write(f"**🛁 Bathrooms:** {house.get('NO OF BATHROOM', '-')}")
                        st.write(f"**📐 Built up size:** {house.get('Built_up_sqft', '-')} sqft")
                        
                        price = house.get('Price_RM')
                        if price is not None and str(price).strip() not in ["", "-", "None"]:
                            try:
                                st.write(f"**💰 Price:** RM {int(float(price)):,}")
                            except ValueError:
                                st.write(f"**💰 Price:** RM {price}")
                        else:
                            st.write("**💰 Price:** RM - (Contact for quotation)")
                    with c2:
                        img_path = get_image_path(house.get('3D_Folder_Path', ''))
                        if img_path:
                            st.image(img_path, use_container_width=True)
                        else:
                            st.info("Floor plan graphic unavailable.")

# --- 5. CHAT LOGIC ---
if user_input := st.chat_input("Type your house requirements here..."):
    st.session_state.messages.append({"role": "user", "content": user_input})
    with st.chat_message("user"):
        st.write(user_input)

    # Clean structured validation format string
    database_context = "Here are the available house plans we can build:\n"
    if all_houses:
        for idx, h in enumerate(all_houses):
            database_context += f"- Plan ID '{h.get('3D_Folder_Path')}': {h.get('NO OF BEDROOM')} Bedrooms, {h.get('NO OF BATHROOM')} Bathrooms, Price: RM {h.get('Price_RM')}, Built-up: {h.get('Built_up_sqft')} sqft. Dimensions: Kitchen {h.get('KITCHEN SIZE')}, Living {h.get('LIVING AREA')}.\n"
    else:
        database_context += "The house plans database is currently completely empty or unreachable.\n"

    system_prompt = (
        "You are an expert house plan consultant chatbot. Your task is to analyze the user's message, "
        "look at the provided house plans database context, and recommend the best plan options. "
        "Be conversational, professional, and friendly. Speak in English or Malay depending on how the user greets you. "
        "CRITICAL RULE: If you recommend any specific plans from the data array, you MUST list their exact plan IDs (e.g., 3D_Models/P01, 3D_Models/P04) at the very bottom of your message "
        "enclosed inside square brackets like this: [MATCHES: 3D_Models/P01, 3D_Models/P04]. If nothing matches or database is empty, do not include the brackets."
    )

    with st.chat_message("assistant"):
        with st.spinner("Analyzing requirements..."):
            try:
                full_prompt = f"{system_prompt}\n\nDatabase Context:\n{database_context}\n\nUser request: {user_input}"
                headers = {'Content-Type': 'application/json'}
                payload = {"contents": [{"parts": [{"text": full_prompt}]}]}
                
                # --- AUTOMATIC ENDPOINT FALLBACK LOOP ---
                models_to_try = ["gemini-1.5-flash", "gemini-2.5-flash"]
                response = None
                response_json = {}
                
                for model_name in models_to_try:
                    url = f"https://generativelanguage.googleapis.com/v1beta/models/{model_name}:generateContent?key={GEMINI_API_KEY}"
                    response = requests.post(url, headers=headers, json=payload)
                    response_json = response.json()
                    
                    if response.status_code == 200:
                        break
                
                if response and response.status_code == 200:
                    ai_text = response_json['candidates'][0]['content']['parts'][0]['text']
                    st.write(ai_text)
                    
                    matched_plans_to_display = []
                    if "[MATCHES:" in ai_text:
                        parsed_section = ai_text.split("[MATCHES:")[-1].replace("]", "").strip()
                        mentioned_ids = [item.strip() for item in parsed_section.split(",")]
                        
                        for house in all_houses:
                            if str(house.get('3D_Folder_Path')).strip() in mentioned_ids:
                                matched_plans_to_display.append(house)
                    
                    new_message = {"role": "assistant", "content": ai_text}
                    if matched_plans_to_display:
                        new_message["matched_plans"] = matched_plans_to_display
                        
                        for house in matched_plans_to_display:
                            with st.expander(f"📋 View Blueprint Details for {house.get('3D_Folder_Path')}"):
                                c1, c2 = st.columns([1, 1])
                                with c1:
                                    st.write(f"**🛏️ Bedrooms:** {house.get('NO OF BEDROOM', '-')}")
                                    st.write(f"**🛁 Bathrooms:** {house.get('NO OF BATHROOM', '-')}")
                                    st.write(f"**📐 Built up size:** {house.get('Built_up_sqft', '-')} sqft")
                                    
                                    price = house.get('Price_RM')
                                    if price is not None and str(price).strip() not in ["", "-", "None"]:
                                        try:
                                            st.write(f"**💰 Price:** RM {int(float(price)):,}")
                                        except ValueError:
                                            st.write(f"**💰 Price:** RM {price}")
                                    else:
                                        st.write("**💰 Price:** RM -")
                                with c2:
                                    img_path = get_image_path(house.get('3D_Folder_Path', ''))
                                    if img_path:
                                        st.image(img_path, use_container_width=True)
                                    else:
                                        st.info("Floor plan graphic unavailable.")
                                        
                    st.session_state.messages.append(new_message)
                else:
                    st.error(f"API Error ({response.status_code if response else 'No Response'}): {response_json.get('error', {}).get('message', 'Models could not be requested.')}")
                    
            except Exception as e:
                st.error(f"Application Error: {e}")
